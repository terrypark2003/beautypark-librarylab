#!/usr/bin/env python3
"""
뷰티파크 '디자인 작업 요청서' 엑셀 → 구조화 JSON 추출기.

업체가 매월 보내는 엑셀(시트명 'YYYY.M')을 읽어, 이벤트 그룹/시술명/가격/강조항목/
월초·월중순 카피를 그대로 뽑아낸다. 사람이 손으로 옮겨 적다 생기는 시술명·가격 오기를
원천 차단하는 것이 목적(과거 최대 마찰요인).

사용법:
  python3 tools/parse_request.py assets/event-requests/2026-07-design-request.xlsx 2026.7
"""
import json
import sys
import openpyxl


def won(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except ValueError:
        return None


def parse_sheet(path, sheet):
    wb = wb_styles = None
    wb = openpyxl.load_workbook(path, data_only=True)
    wb_styles = openpyxl.load_workbook(path)
    ws, wss = wb[sheet], wb_styles[sheet]

    # 가격 컬럼 라벨(1행)과 시술명/타이틀 컬럼 위치는 고정 레이아웃 기준
    title = ws["A1"].value
    price_cols = {
        "정상가": ws["C1"].value,
        "이벤트가": ws["D1"].value,
        "원내확인용": ws["E1"].value,
    }
    deliverables = {
        ws["F3"].value: ws["F4"].value,   # 플친 와이드: 캐러셀(N개)
        ws["G3"].value: ws["G4"].value,   # 플친 리스트: 캐러셀(N개)
        ws["H3"].value: None,             # 인스타(2개)
    }

    groups, current = [], None
    emphasis = None
    for r in range(5, ws.max_row + 1):
        a = ws[f"A{r}"].value
        b = ws[f"B{r}"].value
        if a and str(a).strip().startswith("월 초 강조") or (a and "월초" in str(a)):
            emphasis = str(a).strip()
            continue
        if a and str(a).strip():  # 새 이벤트 그룹 타이틀
            current = {"group": str(a).strip(), "items": []}
            groups.append(current)
        if b and str(b).strip():
            featured = bool(wss[f"B{r}"].font and wss[f"B{r}"].font.bold)
            item = {
                "name": str(b).strip(),
                "정상가": won(ws[f"C{r}"].value),
                "이벤트가": won(ws[f"D{r}"].value),
                "이벤트가_VAT포함": won(ws[f"E{r}"].value),
                "featured": featured,
            }
            if current is None:
                current = {"group": "(미분류)", "items": []}
                groups.append(current)
            current["items"].append(item)

    return {
        "title": title,
        "sheet": sheet,
        "price_columns": price_cols,
        "deliverables": deliverables,
        "emphasis": emphasis,
        "event_groups": [g for g in groups if g["items"]],
    }


if __name__ == "__main__":
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    if sheet is None:
        sheet = openpyxl.load_workbook(path, read_only=True).sheetnames[0]
    data = parse_sheet(path, sheet)
    print(json.dumps(data, ensure_ascii=False, indent=2))
