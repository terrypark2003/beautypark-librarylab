#!/usr/bin/env python3
"""월별 디자인 작업 요청서(엑셀) 생성기.

원내 요청서 서식(A=타이틀 세로병합 · B=시술명 · C=정상가(부가세 전) · D=이벤트가(부가세 전) ·
E=D×1.1 · I=비고)을 그대로 만들어, 광고 에이전시 전달본으로 쓸 수 있게 한다.

사용법:
    python3 tools/build_request_xlsx.py tools/plans/2026-10.json "10월 디자인 작업 요청서.xlsx"

계획 JSON 형식:
    {"sheet": "2026.10", "title": "10월 디자인 작업 요청서",
     "emphasis": "월 초 강조 : …\n월 중순 강조 : …",
     "groups": [{"group": "타이틀", "afterEmphasis": false,
                 "items": [{"name": "...", "normal": 435000, "event": 330000, "note": "..."}]}]}

`afterEmphasis: true`인 그룹(런칭 등 장기 이벤트)은 월초/중순 강조 블록 아래에 배치된다.
`sortItemsByPrice: true`(계획 최상위)면 각 그룹의 시술을 이벤트가 오름차순으로 나열한다 —
`옵션가)`가 들어간 줄은 맨 아래로 보내고, 같은 가격·가격 미정은 원래 순서를 유지한다(원내 검수 요청 관례).
가격 규칙은 CLAUDE.md 참조 — 정상가는 최신 수가표 PDF가 유일한 진실원천이며
n회권 정상가는 반드시 1회가×n으로 검산할 것.
"""
import json
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YELLOW = PatternFill("solid", fgColor="FFFF00")
FONT = "맑은 고딕"


def font(size=10, bold=False):
    return Font(name=FONT, size=size, bold=bold)


def write_header(ws, title):
    ws.merge_cells("A1:B2")
    ws["A1"] = title
    ws["A1"].font = font(14, True)
    ws["A1"].alignment = Alignment("center", "center")
    for col, text in (
        ("C", "정상가\n(부가세 전)"),
        ("D", "이벤트가\n(부가세 전)"),
        ("E", "이벤트가 부가세 포함\n(원내확인용)"),
    ):
        ws.merge_cells(f"{col}1:{col}2")
        ws[f"{col}1"] = text
        ws[f"{col}1"].font = font(9, True)
        ws[f"{col}1"].alignment = Alignment("center", "center", wrap_text=True)
    ws.merge_cells("A3:B3")
    for addr, text in (
        ("F3", "플친 와이드"), ("G3", "플친 리스트"), ("H3", "인스타(2개)"), ("I3", "비고"),
        ("F4", "캐러셀(3개)"), ("G4", "캐러셀(5개)"), ("A4", "타이틀"), ("B4", "시술명"),
    ):
        ws[addr] = text
        ws[addr].font = font(10, True)
        ws[addr].alignment = Alignment("center", "center")
    for col in ("H", "I"):
        ws.merge_cells(f"{col}3:{col}4")
    for col in ("A", "B"):
        ws[f"{col}4"].fill = YELLOW
    for row in range(1, 5):
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = BORDER


def sort_items(items):
    """이벤트가 오름차순. 옵션가 줄은 맨 아래, 가격 미정은 그 다음. 안정 정렬이라 동가는 원래 순서."""
    def key(item):
        is_option = "옵션가)" in item["name"]
        event = item.get("event")
        return (is_option, event is None, event if event is not None else 0)
    return sorted(items, key=key)


def write_groups(ws, groups, row, sort_by_price=False):
    for group in groups:
        start = row
        items = sort_items(group["items"]) if sort_by_price else group["items"]
        for item in items:
            event = item["event"]
            ws.cell(row=row, column=2, value=item["name"])
            if item.get("normal"):
                ws.cell(row=row, column=3, value=item["normal"])
            if event is not None:
                ws.cell(row=row, column=4, value=event)
                ws.cell(row=row, column=5, value=round(event * 1.1))
            if item.get("note"):
                ws.cell(row=row, column=9, value=item["note"])
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = BORDER
                cell.font = font(10)
                if col in (3, 4, 5):
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment("right", "center")
                elif col == 2:
                    cell.alignment = Alignment("left", "center", wrap_text=True)
                elif col == 9:
                    cell.font = font(8)
                    cell.alignment = Alignment("left", "center", wrap_text=True)
            row += 1
        title = ws.cell(row=start, column=1, value=group["group"])
        title.font = font(10, True)
        title.alignment = Alignment("center", "center", wrap_text=True)
        if len(group["items"]) > 1:
            ws.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
    return row


def write_emphasis(ws, text, row):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font(10, True)
    cell.alignment = Alignment("center", "center", wrap_text=True)
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 34


def build(plan, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = plan["sheet"]
    write_header(ws, plan["title"])

    monthly = [g for g in plan["groups"] if not g.get("afterEmphasis")]
    trailing = [g for g in plan["groups"] if g.get("afterEmphasis")]
    by_price = bool(plan.get("sortItemsByPrice"))
    row = write_groups(ws, monthly, 5, by_price)
    if plan.get("emphasis"):
        write_emphasis(ws, plan["emphasis"], row)
        row += 2
    if trailing:
        write_groups(ws, trailing, row, by_price)

    for col, width in (
        ("A", 34), ("B", 60), ("C", 13), ("D", 13),
        ("E", 15), ("F", 11), ("G", 11), ("H", 11), ("I", 46),
    ):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"
    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    with open(sys.argv[1], encoding="utf-8") as fh:
        plan = json.load(fh)
    print("saved:", build(plan, sys.argv[2]))


if __name__ == "__main__":
    main()
